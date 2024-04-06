import math
import os
import sys
import random
import copy
from enum import Enum
from single_engine_predict_intralayer import *
from mesh_hetero import *
from matplotlib import pyplot as plt
from config import *
import openpyxl
import argparse
from basicParam_noc_nop import *
from GaEncode import *
from gaTest_noc_nop import *

def randomTest(CodeGen, Parser, iterTimes, spatial_parallel_list, memory_param, NoC_param, all_sim_node_num , if_multicast, excel_filename, workload_name, i_act_enough, weight_enough, fuse_par_num, fuse_tag, objective="edp", flag = "ours", io_die_tag = 1):

    degrade_ratio_list = []
    excel_datas = []

    best_fitness = 0
    fitness_list = []
    best_fitness_list = []
    best_sample = {}
    best_neu_needed = {}
    best_pkt_needed = {}

    energy_list_d = {}
    latency_list_d = {}
    sp_id = 0

    iters_per_sp = math.ceil(iterTimes / len(spatial_parallel_list))
    for spatial_parallel in spatial_parallel_list:
        sp_energy_list = []
        sp_latency_list = []
        CodeGen.setSpatialParallel(spatial_parallel)
        print("SpatialParallel is {} ----------".format(spatial_parallel))
        for i in range(iters_per_sp):
            print("iterTime----({}, {}, {})".format(i, iters_per_sp, iterTimes))
            # --- 生成个代，与个体解码 --- #
            code = CodeGen.getCode()
            code_c = CodeGen.codeChange(code)
            for_list, act_wgt_dict, out_dict, parallel_dim_list, partition_list, code = Parser.decode(code_c)
            # --- 计算适应度 --- #
            delay, degrade_ratio, degrade_ratio_dict, pkt_needed, neu_needed, compuation_cycles, runtime_list,cp_list,utilization_ratio_list, energy_dram_list, energy_L2_list, energy_L1_list, energy_die2die, energy_MAC, energy_psum_list, delay_psum, worstlinks = \
                calFitness(for_list, act_wgt_dict, out_dict, parallel_dim_list, partition_list, CodeGen.network_param, Parser.HW_param, memory_param, NoC_param, if_multicast, i_act_enough, 0, weight_enough, fuse_par_num, fuse_tag, flag = flag, io_die_tag = io_die_tag)
            
            #---比较适应度，并记录相关变量---
            e_mem = sum(energy_dram_list)+sum(energy_L2_list)+sum(energy_L1_list)
            e_sum = e_mem + energy_die2die+energy_MAC + energy_psum_list[2]
            edp = (delay + delay_psum) * e_sum  /(PE_freq * freq_1G) # pJ*s
            sp_energy_list.append(e_sum)
            sp_latency_list.append(delay + delay_psum)

            if objective == "edp":
                fitness = edp
            elif objective == "energy":
                fitness = e_sum
            elif objective == "latency":
                fitness = delay + delay_psum
            if best_fitness == 0 or fitness < best_fitness:
                best_fitness = fitness
                best_compuation_cycles = compuation_cycles
                best_degrade_ratio = degrade_ratio
                best_degrade_ratio_dict = copy.deepcopy(degrade_ratio_dict)
                best_code = code
                best_sample["for_list"] 			= copy.deepcopy(for_list)
                best_sample["act_wgt_dict"] 		= copy.deepcopy(act_wgt_dict)
                best_sample["out_dict"] 			= copy.deepcopy(out_dict)
                best_sample["parallel_dim_list"] 	= copy.deepcopy(parallel_dim_list)
                best_sample["partition_list"] 		= copy.deepcopy(partition_list)
                best_sample["fitness"] 				= fitness
                best_sample["edp"] 					= edp
                best_sample["energy"] 				= e_sum
                best_sample["delay"] 				= delay + delay_psum
                best_sample["compuation_cycles"] 	= compuation_cycles
                best_sample["degrade_ratio"] 		= degrade_ratio
                best_sample["code"] 				= code
                best_pkt_needed				= copy.deepcopy(pkt_needed)
                best_neu_needed				= copy.deepcopy(neu_needed)
            fitness_list.append(fitness)
            best_fitness_list.append(best_fitness)
            degrade_ratio_list.append (degrade_ratio_dict)

            excel_datas.append([str(spatial_parallel), i, fitness, degrade_ratio, str(for_list[0]), \
                parallel_dim_list[0][0],parallel_dim_list[0][1],parallel_dim_list[0][2],parallel_dim_list[0][3], \
                parallel_dim_list[1][0],parallel_dim_list[1][1],parallel_dim_list[1][2], parallel_dim_list[1][3],\
                parallel_dim_list[0][0]*parallel_dim_list[1][0], \
                parallel_dim_list[0][1]*parallel_dim_list[1][1], \
                parallel_dim_list[0][3]*parallel_dim_list[1][3], \
                parallel_dim_list[0][0]*parallel_dim_list[1][0]*parallel_dim_list[0][1]*parallel_dim_list[1][1], \
                str(partition_list), runtime_list[0], runtime_list[1], runtime_list[2],  \
                runtime_list[3], runtime_list[4], runtime_list[5], runtime_list[6],\
                cp_list[0], cp_list[1], cp_list[2],cp_list[3], cp_list[4], cp_list[5] ,\
                utilization_ratio_list[0], utilization_ratio_list[1], utilization_ratio_list[2],utilization_ratio_list[3], utilization_ratio_list[4], utilization_ratio_list[5], \
                energy_dram_list[0], energy_dram_list[1], energy_dram_list[2], energy_dram_list[3], \
                energy_L2_list[0], energy_L2_list[1], energy_L2_list[2], energy_L2_list[3], \
                energy_L1_list[0], energy_L1_list[1], energy_L1_list[2], energy_psum_list[0], energy_psum_list[1], \
                sum(energy_dram_list), sum(energy_L2_list), sum(energy_L1_list), energy_die2die, energy_MAC, energy_psum_list[2], e_mem, e_sum , delay, edp, str(worstlinks), str(code) ])
            
            print("fitness_min = {}, compuation_cycles = {}, degrade_ratio = {}".format(best_fitness, best_compuation_cycles, str(best_degrade_ratio_dict)))
            
        energy_list_d[sp_id] = sp_energy_list
        latency_list_d[sp_id] = sp_latency_list
        sp_id += 1
    #---生成task file
    task_repeat_nums = createTaskFile(workload_name, best_pkt_needed, best_neu_needed, best_sample["for_list"], best_sample["act_wgt_dict"], best_sample["out_dict"], best_sample["parallel_dim_list"], best_sample["partition_list"],CodeGen.network_param, Parser.HW_param, memory_param, NoC_param, all_sim_node_num, if_multicast)

    #---记录方案的性能指标
    if excel_filename != None:
        workbook = openpyxl.Workbook()
        sheet = workbook.get_sheet_by_name('Sheet') 
        # 写入标题
        column_tite = ["spatial parallel", "index","fitness","degrade_ratio", "dataflow", \
            "PP2","PQ2","PC2","PK2","PP3","PQ3","PC3","PK3","PP","PQ","PKtotal","PPPQtotal", \
            "partition_list",\
            "runtimeP","runtimeQ", "runtimeC", "runtimeK", "runtimeChipNum", "runtimeCoreNum", "runtime_calNum",\
            "ol1_cp_id","al1_cp_id","wl1_cp_id","ol2_cp_id","al2_cp_id","wl2_cp_id", \
            "ol1_util","al1_util","wl1_util","ol2_util","al2_util","wl2_util", \
            "e_wr_opt_dram", "e_rd_opt_dram", "e_rd_wgt_dram", "e_rd_act_dram", \
            "e_wr_opt_L2", "e_rd_opt_L2", "e_rd_wgt_L2", "e_rd_act_L2", \
            "e_rd_wgt_L1", "e_rd_act_L1", "e_rd_opt_L1", "e_psum_d2d", "e_psum_dram", \
            "e_dram", "e_L2", "e_L1", "e_die2die", "e_MAC", "e_psum","e_mem",  "e_sum", "delay", "EDP pJ*s", "worstlinks", "code"]
        for col,column in enumerate(column_tite):
            sheet.cell(1, col+1, column)
        # 写入每一行
        for row, data in enumerate(excel_datas):
            for col, column_data in enumerate(data):
                sheet.cell(row+2, col+1, column_data)

        workbook.save(excel_filename)
    return best_sample, best_degrade_ratio_dict, latency_list_d, energy_list_d, task_repeat_nums


def getLayerParam(app_name):
    abs_path = abs_path = os.path.dirname(os.path.abspath(__file__))
    f = open(abs_path + "/nn_input_noc_nop/" + app_name + ".txt")

    print("network model ----- " + app_name + " -------------")

    lines = f.readlines()
    layer_dict = {}
    layer_name_list = []
    layer_num = 0
    for line in lines:
        if line.startswith("#") or line.startswith("*"):
            pass
        else:
            line = line.replace("\n","")
            line_item = line.split(" ")
            layer_name = line_item[0] + "-" + str(layer_num)
            H = int(line_item[1])
            M = int(line_item[2])
            P = int(line_item[8])
            Q = int(line_item[9])
            C = int(line_item[3])
            K = int(line_item[7])
            R = int(line_item[4])
            S = int(line_item[4])
            stride = int(line_item[5])
            padding = int(line_item[6])
            layer = {"P":P,"Q":Q,"C":C,"K":K,"R":R,"S":S, "stride":stride, "padding":padding, "iact_num":H*M*C, "oact_num":P*Q*K, "w_num":K*C*R*S}
            layer_dict[layer_num] = layer
            layer_name_list.append(layer_name)
            layer_num += 1
    
    layer_list = []
    layer_dict_unique = {}
    layer_name_dict = {}
    layer_num -= 1
    for i, layer in layer_dict.items():
        if i == 0 and i == layer_num:
            partition_tag = 0
            weight_behind = 0
            weight_above = 0
        elif i == 0:
            partition_tag = 0
            weight_behind = layer_dict[i+1]["w_num"]
            weight_above = 0
        elif i == layer_num:
            partition_tag = 2
            weight_behind = 0
            weight_above = layer_dict[i-1]["w_num"]
        else:
            partition_tag = 1
            weight_behind = layer_dict[i+1]["w_num"]
            weight_above = layer_dict[i-1]["w_num"]
        
        if i == layer_num:
            partition_add = 0
        else:
            stride_behind = layer_dict[i+1]["stride"]
            padding_behind = layer_dict[i+1]["padding"]
            R_behind = layer_dict[i+1]["R"]
            partition_add = R_behind - stride_behind - padding_behind
        
        layer["partition"] = [partition_tag, partition_add]
        layer["w_num_behind"] = weight_behind
        layer["w_num_above"] = weight_above
        layer_name = layer_name_list[i]

        print(str(layer_name) + " : " + str(layer))

        if layer not in layer_list:
            layer_dict_unique[layer_name] = layer
            layer_list.append(layer)
            layer_name_dict[layer_name] = layer_name
        else:
            for layer_name_1, layer_1 in layer_dict_unique.items():
                if layer == layer_1:
                    layer_name_same = layer_name_1
                    break
            layer_name_dict[layer_name] = layer_name_same
    
    for i in range(len(layer_name_list)):
        layer_name = layer_name_list[i]
        if layer_name in layer_dict_unique:
            if i < len(layer_name_list)-1:
                i_t = i + 1
                layer_name_tail = layer_name
                while layer_name_tail in layer_name_list[0:i+1]:
                    layer_name_behind = layer_name_list[i_t]
                    layer_name_tail = layer_name_dict[layer_name_behind]
                    i_t += 1
                layer_dict_unique[layer_name]["layer_name_tail"] = layer_name_tail
            else:
                layer_dict_unique[layer_name]["layer_name_tail"] = None

    f.close()
    
    return layer_dict_unique, layer_name_dict

def getLayerParam_pre(app_name):
    layer_dict = {}
    input_activation_num = {}
    layer_name_list = []
    abs_path = abs_path = os.path.dirname(os.path.abspath(__file__))
    f = open(abs_path + "/nn_input_noc_nop/" + app_name + ".txt")

    print("network model ----- " + app_name + " -------------")

    lines = f.readlines()
    for line in lines:
        if line.startswith("#"):
            pass
        elif line.startswith("*"):
            line_item = line.split(" ")
            for i in line_item:
                if i == "*" or i == "end":
                    pass
                else:
                    layer_name_list.append(i)
        else:
            line = line.replace("\n","")
            line_item = line.split(" ")
            layer_name = line_item[0]
            if layer_name in layer_name_list:
                H = int(line_item[1])
                M = int(line_item[2])
                P = int(line_item[8])
                Q = int(line_item[9])
                C = int(line_item[3])
                K = int(line_item[7])
                R = int(line_item[4])
                S = int(line_item[4])
                stride = int(line_item[5])
                partition = str(line_item[11])
                layer_dict[layer_name] = {"P":P,"Q":Q,"C":C,"K":K,"R":R,"S":S, "stride":stride, "partition": partition}
                input_activation_num[layer_name] = H * M * C
                print(str(layer_name) + " : " + str(layer_dict[layer_name]))
    f.close()
    return layer_dict, input_activation_num

def getLayerParamForMulti(layer_dict, mem_size_dict):
    O_mem = mem_size_dict["O"]
    W_mem = mem_size_dict["W"]

    tail_layer_dict = {}
    head_layer_dict = {}
    for layer_name, layer_i in layer_dict.items():
        layer = copy.deepcopy(layer_i)
        partition_tag, partition_add = layer["partition"]
        iact_num = layer["iact_num"]
        oact_num = layer["oact_num"]
        w_num = layer["w_num"]
        w_num_behind = layer["w_num_behind"]
        layer_name_tail = layer["layer_name_tail"]
        if layer_name_tail == None:
            layer_tail = None
        else:
            layer_tail = copy.deepcopy(layer_dict[layer_name_tail])

        if iact_num > O_mem or partition_tag == 0:
            layer_dict[layer_name]["i_act_enough"] = 0
        else:
            layer_dict[layer_name]["i_act_enough"] = 1
        # --- head layer
        par_useful = 1
        if partition_tag <= 1 and oact_num > O_mem and len(layer_dict) > 1:
            weight_enough = ( (w_num + w_num_behind) <= W_mem )
            P_par = 1
            Q_par = 1
            P = layer["P"]
            Q = layer["Q"]
            K = layer["K"]
            P_tile = P
            Q_tile = Q
            while P_tile * Q_tile * K > O_mem:
                if P_tile == 1 and Q_tile == 1:
                    par_useful = 0
                    break
                    
                if 2 * P_tile > Q_tile or Q_tile == 1:
                    P_par *= 2
                else:
                    Q_par *= 2
                P_tile = min(P, (math.ceil(P/P_par) + partition_add))
                Q_tile = min(Q, (math.ceil(Q/Q_par) + partition_add))
                
            if par_useful == 1:
                layer["P"] = P_tile
                layer["Q"] = Q_tile
                layer["P_par"] = P_par
                layer["Q_par"] = Q_par
                layer["weight_enough"] = weight_enough
                layer["i_act_enough"] = layer_dict[layer_name]["i_act_enough"]
                head_layer_dict[layer_name] = layer

                P_t, Q_t = layer_tail["P"], layer_tail["Q"]
                layer_tail["P"], layer_tail["Q"] = math.ceil(P_t / P_par), math.ceil(Q_t / Q_par)
                layer_tail["P_par"], layer_tail["Q_par"] = P_par, Q_par
                layer_tail["weight_enough"] = weight_enough
                layer_tail["i_act_enough"] = 1
                tail_layer_dict[layer_name_tail] = layer_tail

    return layer_dict, head_layer_dict, tail_layer_dict

def getLayerParamForMulti_pre(layer_dict, input_activation_num, partiton_size_list):
    tail_layer_dict = {}
    head_layer_dict = {}
    tail_iact_num_dict = {}
    head_iact_num_dict = {}
    for layer_name in layer_dict:
        partition = layer_dict[layer_name]["partition"]
        partiton_list = partition.split("-")
        partition_tag = int(partiton_list[0])
        partition_offset = int(partiton_list[1])

        tail_dict = copy.deepcopy(layer_dict[layer_name])
        tail_dict["P"] = math.ceil(tail_dict["P"] / partiton_size_list["P"])
        tail_dict["Q"] = math.ceil(tail_dict["Q"] / partiton_size_list["Q"])

        head_dict = copy.deepcopy(tail_dict)
        head_dict["P"] = head_dict["P"] + partition_offset
        head_dict["Q"] = head_dict["Q"] + partition_offset

        iact_num = math.ceil(input_activation_num[layer_name] / (partiton_size_list["P"] * partiton_size_list["Q"]) )
        
        if partition_tag == 3:
            pass
        elif partition_tag == 0:
            head_layer_dict[layer_name] = head_dict
            head_iact_num_dict[layer_name] = input_activation_num[layer_name]
        elif partition_tag == 1:
            head_layer_dict[layer_name] = head_dict
            head_iact_num_dict[layer_name] = input_activation_num[layer_name]
            tail_layer_dict[layer_name] = tail_dict
            tail_iact_num_dict[layer_name] = iact_num
        elif partition_tag == 2:
            tail_layer_dict[layer_name] = tail_dict
            tail_iact_num_dict[layer_name] = iact_num

    return tail_layer_dict, tail_iact_num_dict, head_layer_dict, head_iact_num_dict

def getSPPartitonList(num, sp_dim, sp_type, TH = 10):
    list = []
    sp_init = {"P":1, "Q":1, "C":1, "K":1, "R":1, "S":1}
    gen_num = 0
    iter_num = 0
    if sp_type == 1:
        #--- single
        for dim_id in sp_dim:
            dim = dim_list[dim_id]
            sp_dict = copy.deepcopy(sp_init)
            sp_dict[dim] = num
            list.append(sp_dict)
        return list
    elif sp_type == 3:
        # 均匀
        assert(len(sp_dim) == 2)
        num_half = int(pow(num, 0.5))
        #assert(num_half * num_half == num)
        num_list = []
        if num_half * num_half == num:
            num_list.append([num_half, num_half])
        else:
            num1 = num_half - 1
            if num1 == 0:
                num1 = 1
            while num % num1 != 0:
                num1 -= 1
            num2 = int(num / num1)
            num_list.append([num1, num2])
            num_list.append([num2, num1])
        dim1 = dim_list[sp_dim[0]]
        dim2 = dim_list[sp_dim[1]]
        for num_l in num_list:
            sp_dict = copy.deepcopy(sp_init)
            sp_dict[dim1] *= num_l[0]
            sp_dict[dim2] *= num_l[1]
            list.append(sp_dict)
        return list
    elif sp_type == 0:
        #--- no limit
        if len(sp_dim) == 1:
            dim = dim_list[sp_dim[0]]
            sp_dict = copy.deepcopy(sp_init)
            sp_dict[dim] = num
            list.append(sp_dict)
            return list
        else:
            while gen_num < TH:
                ran1 = random.randint(0, len(sp_dim)-1)
                ran2 = random.randint(0, len(sp_dim)-1)
                dim1 = dim_list[sp_dim[ran1]]
                dim2 = dim_list[sp_dim[ran2]]
                [num1, num2, num3] = setPartition_1(num, 2)
                sp_dict = copy.deepcopy(sp_init)
                sp_dict[dim1] *= num1
                sp_dict[dim2] *= num2
                if sp_dict not in list:
                    gen_num += 1
                    list.append(sp_dict)
                
                iter_num +=1
                if iter_num > 1000:
                    break
            return list
    elif sp_type == 2:
        #--- hybrid
        assert(len(sp_dim) >= 2)
        while gen_num < TH:
            id_list = list(range(len(sp_dim)))
            random.shuffle(id_list)
            dim1 = dim_list[sp_dim[id_list[0]]]
            dim2 = dim_list[sp_dim[id_list[1]]]
            [num1, num2, num3] = setPartition_1(num, 2)
            sp_dict = copy.deepcopy(sp_init)
            sp_dict[dim1] *= num1
            sp_dict[dim2] *= num2
            if sp_dict not in list and num1 != 1 and num2 != 1:
                gen_num += 1
                list.append(sp_dict)
            
            iter_num +=1
            if iter_num > 1000:
                break
        return list

def getSpatialParallel(HW_param, chiplet_parallel, PE_parallel, numTH = 20):

    spatial_parallel_init = {"pe":{"P":1, "Q":1, "C":1, "K":1, "R":1, "S":1}, "chiplet":{"P":1, "Q":1, "C":1, "K":1, "R":1, "S":1}, "package":{"P":1, "Q":1, "C":1, "K":1, "R":1, "S":1}}
    for dim in HW_param["intra_PE"]:
        spatial_parallel_init["pe"][dim] = HW_param["intra_PE"][dim]
    
    chiplet_H = HW_param["Chiplet"][0]
    chiplet_W = HW_param["Chiplet"][1]
    chiplet_num = chiplet_H * chiplet_W
    PE_H = HW_param["PE"][0]
    PE_W = HW_param["PE"][1]
    PE_num = PE_H * PE_W

    parallel_select, parallel_type = config_parallel_type(chiplet_parallel,PE_parallel)	
    chiplet_sp_dim = parallel_select["Chiplet"]
    chiplet_sp_type = parallel_type["Chiplet"]
    pe_sp_dim = parallel_select["PE"]
    pe_sp_type = parallel_type["PE"]

    chiplet_list = getSPPartitonList(chiplet_num, chiplet_sp_dim, chiplet_sp_type)
    TH_pe = int(numTH/len(chiplet_list))
    pe_list = getSPPartitonList(PE_num, pe_sp_dim, pe_sp_type, TH_pe)

    spatial_parallel_list = []

    for chiplet in chiplet_list:
        for pe in pe_list:
            spatial_parallel = copy.deepcopy(spatial_parallel_init)
            spatial_parallel["chiplet"] = pe
            spatial_parallel["package"] = chiplet
            spatial_parallel_list.append(spatial_parallel)
    
    return spatial_parallel_list

def randomTestScatterPlot(latency_list_d, energy_list_d, layer_name, outDir):
    color_list = ["firebrick", "indianred", "chocolate", "tomato", "sandybrown", "orange", "gold", "goldenrod", "olive", "y", \
                    "olivedrab", "g", "teal", "steelblue", "dodgerblue", "royalblue", "mediumpurple", "darkviolet", \
                    "mediumorchid", "hotpink", "palevioletred", "crimson", "pink"]
    plt.figure("Random Test Energy Latency")
    plt.subplot(2,1,1)
    #plt.xlabel("Latency", fontsize = 10)
    plt.ylabel("Energy", fontsize = 10)
    plt.tick_params(labelsize=8)
    good_point_list = []
    for sp_id in latency_list_d:
        latency_list = latency_list_d[sp_id]
        energy_list = energy_list_d[sp_id]
        color = color_list[sp_id]
        print("color: ", color)
        for i in range(len(latency_list)):
            plt.scatter(latency_list[i],energy_list[i],s=2,color=color)
            if latency_list[i] < 100000 and energy_list[i] < 1000000000:
                good_point_list.append([latency_list[i], energy_list[i], color])
    plt.xscale('log')
    plt.yscale('log')

    plt.subplot(2,1,2)
    plt.xlabel("Latency", fontsize = 10)
    plt.ylabel("Energy", fontsize = 10)
    plt.tick_params(labelsize=8)
    for point in good_point_list:
        plt.scatter(point[0],point[1],s=2,color=point[2])
    plt.xscale('log')
    plt.yscale('log')
    plt.tight_layout(pad=1.1)
    plt.savefig(outDir + "/" + layer_name + "_scatter_plot.png", bbox_inches = 'tight')

def randomTest_NoC(mid_result_record_file, nn_name, iterTime, result_dir, save_all_records, record_dir, GaType, HW_param, memory_param, layer_dict, spatial_parallel_list, temporal_level, NoC_param, optimization_objective, layer_name_dict, all_sim_node_num, multi_layer_tag="initial", if_multicast=1, io_die_tag = 1):
    
    best_edp_dict = {}
    best_energy_dict = {}
    best_delay_dict = {}
    best_code_dict = {}
    best_degrade_ratio_dict = {}
    NoC_DR_dict = {}
    L2_to_DRAM_DR_dict = {}
    DRAM_to_L2_DR_dict = {}
    NoP_DR_dict = {}
    best_degrade_ratio_dict_dict = {}
    best_compuation_cycles_dict = {}

    best_sample_dict = {}

    edp_total = 0

    latency_list_dict = {}
    energy_list_dict = {}

    i_act_enough_dict = {}
    layer_par_num_dict = {}
    layer_par_num_detail_dict = {}
    task_repeat_nums_dict = {}

    for layer_name in layer_dict:
        # --- 输出文件名 --- #
        if save_all_records == 1:
            record_filename = record_dir + "/" + layer_name + "_" + multi_layer_tag + ".xlsx"
        else:
            record_filename = None
        layer_n = layer_name.split("-")
        workload_name = "{}_{}".format(nn_name, layer_n[0])

        # --- 初始化参数 --- #
        network_param = layer_dict[layer_name]
        i_act_enough = layer_dict[layer_name]['i_act_enough']
        i_act_enough_dict[layer_name] = i_act_enough
        if "weight_enough" in layer_dict[layer_name]:
            weight_enough = layer_dict[layer_name]["weight_enough"]
            par_num = layer_dict[layer_name]['P_par'] * layer_dict[layer_name]['Q_par']
            layer_par_num_detail_dict[layer_name] = [layer_dict[layer_name]['P_par'], layer_dict[layer_name]['Q_par']]
        else:
            weight_enough = 0
            par_num = 1
            layer_par_num_detail_dict[layer_name] = [1, 1]
        layer_par_num_dict[layer_name] = par_num

        CodeGen = Encoder(GaType, network_param, temporal_level, debug=0)
        Parser = Decoder(temporal_level, HW_param, network_param)

        best_sample, degrade_ratio_dict, latency_list_d, energy_list_d, task_repeat_nums = \
            randomTest(CodeGen, Parser, iterTime, spatial_parallel_list, memory_param, NoC_param, all_sim_node_num, if_multicast, record_filename, workload_name, i_act_enough, weight_enough, par_num, multi_layer_tag, io_die_tag = io_die_tag)
        best_edp_dict[layer_name] 		= best_sample["edp"]
        best_energy_dict[layer_name] 	= best_sample["energy"]
        best_delay_dict[layer_name] 	= best_sample["delay"]
        best_code_dict[layer_name] 		= best_sample["code"]
        best_degrade_ratio_dict[layer_name] = best_sample["degrade_ratio"]
        NoC_DR_dict[layer_name] = degrade_ratio_dict["NoC"]
        L2_to_DRAM_DR_dict[layer_name] = degrade_ratio_dict["L2_to_DRAM"]
        DRAM_to_L2_DR_dict[layer_name] = degrade_ratio_dict["DRAM_to_L2"]
        best_degrade_ratio_dict_dict[layer_name] = degrade_ratio_dict
        best_compuation_cycles_dict[layer_name] = best_sample["compuation_cycles"]
        edp_total += best_sample["edp"]
        latency_list_dict[layer_name] = latency_list_d
        energy_list_dict[layer_name] = energy_list_d

        best_sample_dict[layer_name] = best_sample

        task_repeat_nums_dict[layer_name] = task_repeat_nums

        # randomTestScatterPlot(latency_list_d, energy_list_d, layer_name, result_dir)
    
    file_1 = result_dir + "/final_result_record_" + multi_layer_tag + ".txt"
    f = open(file_1,'w')
    print(best_edp_dict, file=f)
    print(best_energy_dict, file=f)
    print(best_delay_dict, file=f)
    print(best_code_dict, file = f)
    print(best_degrade_ratio_dict, file = f)
    print(NoC_DR_dict, file = f)
    print(L2_to_DRAM_DR_dict, file = f)
    print(DRAM_to_L2_DR_dict, file = f)
    print(best_degrade_ratio_dict_dict, file = f)
    print(best_compuation_cycles_dict, file = f)
    print("edp_total: ", edp_total, file = f)
    f.close()

    file_2 = mid_result_record_file
    f_2 = open(file_2, 'w')
    for layer_name, best_sample in best_sample_dict.items():
        print("layer_name\t{}".format(layer_name), file=f_2)
        for index, value in best_sample.items():
            print("---\t{}\t{}".format(index, value), file=f_2)
        print("task_repeat_nums: ", task_repeat_nums_dict[layer_name], file=f_2)
        print("", file=f_2)
    f_2.close()
    return best_sample_dict


if __name__ == '__main__':
    # 目前只支持我们的架构，nnbaton和simba还没添加
    # todo : support nnbaton and simba
    parser = argparse.ArgumentParser()
    parser.add_argument('--architecture', type=str, default="ours", help='hardware architecture type (ours, nnbaton, simba)')	# simba , nnbaton
    parser.add_argument('--app_name', type=str, default="resnet50", help='NN model name')

    parser.add_argument('--chiplet_num_max', type=int, default=None, help='chiplet_num_max')
    parser.add_argument('--chiplet_num_min', type=int, default=None, help='chiplet_num_min')
    parser.add_argument('--chiplet_parallel', type=str, default="All", help='chiplet level spatial parallel type') # K_stable, P_stable, PK_stable, C_stable, KC_stable
    parser.add_argument('--PE_parallel', type=str, default="All", help='PE level spatial parallel type')
    parser.add_argument('--debug_open', type=int, default=0, help='debug mode (will print info)')
    parser.add_argument('--save_all_records', type=int, default=0, help='save all record')
    parser.add_argument('--layer_fuse_tag', type=int, default=1, help='layer_fuse_tag')
    parser.add_argument('--optimization_objective', type=str, default="edp", help='optimization_objective')
    parser.add_argument('--temporal_level', type=int, default=3, help='temporal level')
    opt = parser.parse_args()

    abs_path = os.path.dirname(os.path.abspath(__file__))
    architecture = opt.architecture
    app_name = opt.app_name
    alg = "random"
    encode_type = 'index'
    dataflow = 'ours'
    chiplet_num_max = opt.chiplet_num_max
    chiplet_num_min = opt.chiplet_num_min
    chiplet_parallel = opt.chiplet_parallel
    PE_parallel = opt.PE_parallel
    debug_open = opt.debug_open
    save_all_records = opt.save_all_records
    layer_fuse_tag = opt.layer_fuse_tag
    optimization_objective = opt.optimization_objective
    temporal_level = opt.temporal_level

    record_outdir = os.path.join(abs_path, "output_record")
    os.makedirs(record_outdir, exist_ok=True)
    record_outdir = os.path.join(record_outdir, architecture + "_" + app_name)
    os.makedirs(record_outdir, exist_ok=True)
    record_outdir = os.path.join(record_outdir, chiplet_parallel + "_and_" + PE_parallel)
    os.makedirs(record_outdir, exist_ok=True)
    record_outdir = os.path.join(record_outdir, alg + "_" + encode_type + "_" + optimization_objective)
    os.makedirs(record_outdir, exist_ok=True)

    # --- 最终结果记录
    debug_final = 1
    if debug_final:
        file_debug_final = app_name + "_result_record.txt"
        if os.path.exists(file_debug_final) == False:
            f_debug_final = open(file_debug_final, 'w')
        else:
            f_debug_final = open(file_debug_final, 'a')
        print("{:-^120}".format(" SETTING "), file = f_debug_final)
        print("architecture={}".format(architecture), file = f_debug_final)
        print("alg={}".format(alg), file = f_debug_final)
        print("encode_type={}".format(encode_type), file = f_debug_final)
        print("chiplet_num=[{} ~ {}]".format(chiplet_num_max, chiplet_num_min), file = f_debug_final)
        print("chiplet_parallel={}, PE_parallel={}".format(chiplet_parallel, PE_parallel), file = f_debug_final)
        print("temporal_level={}".format(temporal_level), file = f_debug_final)
        print("optimization_objective={}".format(optimization_objective), file = f_debug_final)
        print("{:-^120}".format(" RESULT "), file = f_debug_final)
    
    for i in range(chiplet_num_min-1, chiplet_num_max):
        chiplet_num = i+1
        result_outdir = os.path.join(abs_path, "result")
        os.makedirs(result_outdir, exist_ok=True)
        result_outdir = os.path.join(result_outdir, "intraLayer")
        os.makedirs(result_outdir, exist_ok=True)
        result_outdir = os.path.join(result_outdir, architecture + "_" + app_name)
        os.makedirs(result_outdir, exist_ok=True)
        result_outdir = os.path.join(result_outdir, "chiplet_num_"+str(chiplet_num))
        os.makedirs(result_outdir, exist_ok=True)
        result_outdir = os.path.join(result_outdir, alg + "_" + encode_type +"_" +  optimization_objective)
        os.makedirs(result_outdir, exist_ok=True)
        result_outdir = os.path.join(result_outdir, chiplet_parallel + "_and_" + PE_parallel)
        os.makedirs(result_outdir, exist_ok=True)

        mid_result_dir = os.path.join(abs_path, "best_sample_record")
        os.makedirs(mid_result_dir, exist_ok=True)
        mid_result_record = os.path.join(mid_result_dir, "{}_c{}_result.txt".format(app_name, chiplet_num))

        # --- 硬件参数
        HW_param = {"Chiplet":[4,4],"PE":[4,4],"intra_PE":{"C":8,"K":8}}       	
        memory_param = {"OL1":3 ,"OL2":48,"AL1":8,"AL2":48,"WL1":32,"WL2":32}		# mesh mem
        noc_topo = 'Mesh'

        if architecture == "cmesh_DDR4" or architecture == "cmesh_HBM":
            HW_param["maxChiplet"] = [4, 4]
            io_die_tag = 1
        elif architecture == "mesh_DDR4" or architecture == "mesh_HBM":
            HW_param["maxChiplet"] = [4, 4]
            io_die_tag = 0
        elif architecture == "ring_DDR4" or architecture == "ring_HBM":
            HW_param["maxChiplet"] = [4, 2]
            io_die_tag = 0
        else:
            print("Error architecture(), not supported".format(architecture))
            raise NotImplementedError


        if chiplet_num != None:
            HW_param["Chiplet"] = [1, int(chiplet_num)]

        NoC_w = HW_param["PE"][1] + 1
        NOC_NODE_NUM = NoC_w * HW_param["PE"][0]
        NoP_w = HW_param["Chiplet"][1] + 1
        NOP_SIZE = NoP_w * HW_param["Chiplet"][0]
        TOPO_param = {"NoC_w":NoC_w, "NOC_NODE_NUM": NOC_NODE_NUM, "NoP_w": NoP_w, "NOP_SIZE": NOP_SIZE,"nop_scale_ratio": nop_bandwidth/noc_bandwidth}
        
        # --- 生成noc-nop结构图
        NoC_param, all_sim_node_num = construct_noc_nop_topo(TOPO_param["NOC_NODE_NUM"],TOPO_param["NoC_w"], TOPO_param["NOP_SIZE"],TOPO_param["NoP_w"], TOPO_param["nop_scale_ratio"], topology = noc_topo)
        if_multicast = 0

        # --- 神经网络参数
        layer_dict, layer_name_dict = getLayerParam(app_name)
        partiton_size_list = {"P":2, "Q":2}
        chiplet_num = HW_param["Chiplet"][0] * HW_param["Chiplet"][1]
        OL2_mem = memory_param["OL2"]*8*1024/act_wgt_width * chiplet_num
        WL2_mem = memory_param["WL2"]*8*1024/act_wgt_width * chiplet_num
        mem_size_dict = {"O":OL2_mem, "W":WL2_mem}
        layer_dict, head_layer_dict, tail_layer_dict = getLayerParamForMulti(layer_dict, mem_size_dict)
        print("layer_name_dict----------------------------------")
        # print(layer_name_dict)
        layer_num_dict = {}
        for r_name, a_name in layer_name_dict.items():
            if a_name not in layer_num_dict:
                layer_num_dict[a_name] = 0
            layer_num_dict[a_name] += 1
        for name, num in layer_num_dict.items():
            print(name, "\t\t", num)
        print(" ")
        print("layer_dict---------------------------------------")
        for layer_name, layer in layer_dict.items():
            print(layer_name, ": ", layer)
        print(" ")
        print("head_layer_dict----------------------------------")
        for layer_name, layer in head_layer_dict.items():
            print(layer_name, ": ", layer)
        print(" ")
        print("tail_layer_dict----------------------------------")
        for layer_name, layer in tail_layer_dict.items():
            print(layer_name, ": ", layer)
        assert(len(head_layer_dict) == len(tail_layer_dict))
        #if layer_fuse_tag == 1:
        #	tail_layer_dict, tail_iact_num_dict, head_layer_dict, head_iact_num_dict = getLayerParamForMulti(layer_dict, input_activation_num, partiton_size_list)

        # --- 获得空间并行度
        spatial_parallel_list = getSpatialParallel(HW_param, chiplet_parallel, PE_parallel)
        
        # --- 迭代参数
        num_gen = 10
        num_iter = 1
        iterTime = num_gen * num_iter

        best_sample_dict = randomTest_NoC(mid_result_record, "c{}_{}".format(chiplet_num, app_name), iterTime, result_outdir, save_all_records, record_outdir, encode_type, HW_param, memory_param, layer_dict, spatial_parallel_list, temporal_level, NoC_param, optimization_objective, layer_name_dict, all_sim_node_num, if_multicast=if_multicast, io_die_tag=io_die_tag)
        if layer_fuse_tag == 1:
            best_sample_dict = randomTest_NoC(mid_result_record, "c{}_{}".format(chiplet_num, app_name), iterTime, result_outdir, save_all_records, record_outdir, encode_type, HW_param, memory_param, layer_dict, spatial_parallel_list, temporal_level, NoC_param, optimization_objective, layer_name_dict, all_sim_node_num, multi_layer_tag = "headFuse", if_multicast=if_multicast, io_die_tag=io_die_tag)
            best_sample_dict = randomTest_NoC(mid_result_record, "c{}_{}".format(chiplet_num, app_name), iterTime, result_outdir, save_all_records, record_outdir, encode_type, HW_param, memory_param, layer_dict, spatial_parallel_list, temporal_level, NoC_param, optimization_objective, layer_name_dict, all_sim_node_num, multi_layer_tag = "tailFuse", if_multicast=if_multicast, io_die_tag=io_die_tag)

        for layer_name, best_sample in best_sample_dict.items():
            print("{}\t{}".format(layer_name, best_sample["degrade_ratio"] * best_sample["delay"]))
        print("----------------------------------------------------------------")
        print("-------------------------- SIM END -----------------------------")
        print("----------------------------------------------------------------")
