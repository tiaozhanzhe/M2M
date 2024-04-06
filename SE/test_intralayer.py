import math
import os
import sys
import random
import numpy as np
import copy
from enum import Enum
from single_engine_predict_intralayer import *
from mesh_hetero import *
from matplotlib import pyplot as plt
from config import *
import openpyxl
import argparse
from basicParam_noc_nop import *
from GaEncode import *
from gaTest_noc_nop import *

# --- head, tail的判断以及作为头层增加的列数自动化生成  
# --- layer_dict_unique[layer_name] = {}  
# --- layer_name_dict[layer_name] = layer_name_same  
def getLayerParam(app_name):
	abs_path = abs_path = os.path.dirname(os.path.abspath(__file__))
	f = open(abs_path + "/nn_input_noc_nop/" + app_name + ".txt")

	print("network model ----- " + app_name + " -------------")

	lines = f.readlines()
	layer_dict = {}
	layer_name_list = []
	layer_num = 0
	for line in lines:
		if line.startswith("#") or line.startswith("*"):
			pass
		else:
			line = line.replace("\n","")
			line_item = line.split(" ")
			layer_name = line_item[0] + "-" + str(layer_num)
			H = int(line_item[1])
			M = int(line_item[2])
			P = int(line_item[8])
			Q = int(line_item[9])
			C = int(line_item[3])
			K = int(line_item[7])
			R = int(line_item[4])
			S = int(line_item[4])
			stride = int(line_item[5])
			padding = int(line_item[6])
			layer = {"P":P,"Q":Q,"C":C,"K":K,"R":R,"S":S, "stride":stride, "padding":padding, "iact_num":H*M*C, "oact_num":P*Q*K, "w_num":K*C*R*S}
			layer_dict[layer_num] = layer
			layer_name_list.append(layer_name)
			layer_num += 1
	
	layer_list = []
	layer_dict_unique = {}
	layer_name_dict = {}
	layer_num -= 1
	for i, layer in layer_dict.items():
		if i == 0:
			partition_tag = 0
			weight_behind = layer_dict[i+1]["w_num"]
			weight_above = 0
		elif i == layer_num:
			partition_tag = 2
			weight_behind = 0
			weight_above = layer_dict[i-1]["w_num"]
		else:
			partition_tag = 1
			weight_behind = layer_dict[i+1]["w_num"]
			weight_above = layer_dict[i-1]["w_num"]
		
		if i == layer_num:
			partition_add = 0
		else:
			stride_behind = layer_dict[i+1]["stride"]
			padding_behind = layer_dict[i+1]["padding"]
			R_behind = layer_dict[i+1]["R"]
			partition_add = R_behind - stride_behind - padding_behind
		
		layer["partition"] = [partition_tag, partition_add]
		layer["w_num_behind"] = weight_behind
		layer["w_num_above"] = weight_above
		layer_name = layer_name_list[i]

		print(str(layer_name) + " : " + str(layer))

		if layer not in layer_list:
			layer_dict_unique[layer_name] = layer
			layer_list.append(layer)
			layer_name_dict[layer_name] = layer_name
		else:
			for layer_name_1, layer_1 in layer_dict_unique.items():
				if layer == layer_1:
					layer_name_same = layer_name_1
					break
			layer_name_dict[layer_name] = layer_name_same
	
	for i in range(len(layer_name_list)):
		layer_name = layer_name_list[i]
		if layer_name in layer_dict_unique:
			if i < len(layer_name_list)-1:
				i_t = i + 1
				layer_name_tail = layer_name
				while layer_name_tail in layer_name_list[0:i+1]:
					layer_name_behind = layer_name_list[i_t]
					layer_name_tail = layer_name_dict[layer_name_behind]
					i_t += 1
				layer_dict_unique[layer_name]["layer_name_tail"] = layer_name_tail
			else:
				layer_dict_unique[layer_name]["layer_name_tail"] = None

	f.close()
	
	return layer_dict_unique, layer_name_dict

def getLayerParam_pre(app_name):
	layer_dict = {}
	input_activation_num = {}
	layer_name_list = []
	abs_path = abs_path = os.path.dirname(os.path.abspath(__file__))
	f = open(abs_path + "/nn_input_noc_nop/" + app_name + ".txt")

	print("network model ----- " + app_name + " -------------")

	lines = f.readlines()
	for line in lines:
		if line.startswith("#"):
			pass
		elif line.startswith("*"):
			line_item = line.split(" ")
			for i in line_item:
				if i == "*" or i == "end":
					pass
				else:
					layer_name_list.append(i)
		else:
			line = line.replace("\n","")
			line_item = line.split(" ")
			layer_name = line_item[0]
			if layer_name in layer_name_list:
				H = int(line_item[1])
				M = int(line_item[2])
				P = int(line_item[8])
				Q = int(line_item[9])
				C = int(line_item[3])
				K = int(line_item[7])
				R = int(line_item[4])
				S = int(line_item[4])
				stride = int(line_item[5])
				partition = str(line_item[11])
				layer_dict[layer_name] = {"P":P,"Q":Q,"C":C,"K":K,"R":R,"S":S, "stride":stride, "partition": partition}
				input_activation_num[layer_name] = H * M * C
				print(str(layer_name) + " : " + str(layer_dict[layer_name]))
	f.close()
	return layer_dict, input_activation_num

def getLayerParamForMulti(layer_dict, mem_size_dict):
	O_mem = mem_size_dict["O"]
	W_mem = mem_size_dict["W"]

	tail_layer_dict = {}
	head_layer_dict = {}
	for layer_name, layer_i in layer_dict.items():
		layer = copy.deepcopy(layer_i)
		partition_tag, partition_add = layer["partition"]
		iact_num = layer["iact_num"]
		oact_num = layer["oact_num"]
		w_num = layer["w_num"]
		w_num_behind = layer["w_num_behind"]
		layer_name_tail = layer["layer_name_tail"]
		if layer_name_tail == None:
			layer_tail = None
		else:
			layer_tail = copy.deepcopy(layer_dict[layer_name_tail])

		if iact_num > O_mem or partition_tag == 0:
			layer_dict[layer_name]["i_act_enough"] = 0
		else:
			layer_dict[layer_name]["i_act_enough"] = 1
		# --- head layer
		par_useful = 1
		if partition_tag <= 1 and oact_num > O_mem:
			weight_enough = ( (w_num + w_num_behind) <= W_mem )
			P_par = 1
			Q_par = 1
			P = layer["P"]
			Q = layer["Q"]
			K = layer["K"]
			P_tile = P
			Q_tile = Q
			while P_tile * Q_tile * K > O_mem:
				if P_tile == 1 and Q_tile == 1:
					par_useful = 0
					break
					
				if 2 * P_tile > Q_tile or Q_tile == 1:
					P_par *= 2
				else:
					Q_par *= 2
				P_tile = min(P, (math.ceil(P/P_par) + partition_add))
				Q_tile = max(min(Q, (math.ceil(Q/Q_par) + partition_add)), 1)
				print('Q_tile = ', Q_tile)
				print('math.ceil(Q/Q_par) + partition_add = ', math.ceil(Q/Q_par) + partition_add)

			
			if par_useful == 1:
				layer["P"] = P_tile
				layer["Q"] = Q_tile
				layer["P_par"] = P_par
				layer["Q_par"] = Q_par
				layer["weight_enough"] = weight_enough
				layer["i_act_enough"] = layer_dict[layer_name]["i_act_enough"]
				head_layer_dict[layer_name] = layer

				P_t, Q_t = layer_tail["P"], layer_tail["Q"]
				layer_tail["P"], layer_tail["Q"] = math.ceil(P_t / P_par), math.ceil(Q_t / Q_par)
				layer_tail["P_par"], layer_tail["Q_par"] = P_par, Q_par
				layer_tail["weight_enough"] = weight_enough
				layer_tail["i_act_enough"] = 1
				tail_layer_dict[layer_name_tail] = layer_tail

	return layer_dict, head_layer_dict, tail_layer_dict

def getLayerParamForMulti_pre(layer_dict, input_activation_num, partiton_size_list):
	tail_layer_dict = {}
	head_layer_dict = {}
	tail_iact_num_dict = {}
	head_iact_num_dict = {}
	for layer_name in layer_dict:
		partition = layer_dict[layer_name]["partition"]
		partiton_list = partition.split("-")
		partition_tag = int(partiton_list[0])
		partition_offset = int(partiton_list[1])

		tail_dict = copy.deepcopy(layer_dict[layer_name])
		tail_dict["P"] = math.ceil(tail_dict["P"] / partiton_size_list["P"])
		tail_dict["Q"] = math.ceil(tail_dict["Q"] / partiton_size_list["Q"])

		head_dict = copy.deepcopy(tail_dict)
		head_dict["P"] = head_dict["P"] + partition_offset
		head_dict["Q"] = head_dict["Q"] + partition_offset

		iact_num = math.ceil(input_activation_num[layer_name] / (partiton_size_list["P"] * partiton_size_list["Q"]) )
		
		if partition_tag == 3:
			pass
		elif partition_tag == 0:
			head_layer_dict[layer_name] = head_dict
			head_iact_num_dict[layer_name] = input_activation_num[layer_name]
		elif partition_tag == 1:
			head_layer_dict[layer_name] = head_dict
			head_iact_num_dict[layer_name] = input_activation_num[layer_name]
			tail_layer_dict[layer_name] = tail_dict
			tail_iact_num_dict[layer_name] = iact_num
		elif partition_tag == 2:
			tail_layer_dict[layer_name] = tail_dict
			tail_iact_num_dict[layer_name] = iact_num

	return tail_layer_dict, tail_iact_num_dict, head_layer_dict, head_iact_num_dict

def getSPPartitonList(num, sp_dim, sp_type, TH = 10):
	list = []
	sp_init = {"P":1, "Q":1, "C":1, "K":1, "R":1, "S":1}
	gen_num = 0
	iter_num = 0
	if sp_type == 1:
		#--- single
		for dim_id in sp_dim:
			dim = dim_list[dim_id]
			sp_dict = copy.deepcopy(sp_init)
			sp_dict[dim] = num
			list.append(sp_dict)
		return list
	elif sp_type == 3:
		# 均匀
		assert(len(sp_dim) == 2)
		num_half = int(pow(num, 0.5))
		#assert(num_half * num_half == num)
		num_list = []
		if num_half * num_half == num:
			num_list.append([num_half, num_half])
		else:
			num1 = num_half - 1
			if num1 == 0:
				num1 = 1
			while num % num1 != 0:
				num1 -= 1
			num2 = int(num / num1)
			num_list.append([num1, num2])
			num_list.append([num2, num1])
		dim1 = dim_list[sp_dim[0]]
		dim2 = dim_list[sp_dim[1]]
		for num_l in num_list:
			sp_dict = copy.deepcopy(sp_init)
			sp_dict[dim1] *= num_l[0]
			sp_dict[dim2] *= num_l[1]
			list.append(sp_dict)
		return list
	elif sp_type == 0:
		#--- no limit
		if len(sp_dim) == 1:
			dim = dim_list[sp_dim[0]]
			sp_dict = copy.deepcopy(sp_init)
			sp_dict[dim] = num
			list.append(sp_dict)
			return list
		else:
			while gen_num < TH:
				ran1 = random.randint(0, len(sp_dim)-1)
				ran2 = random.randint(0, len(sp_dim)-1)
				dim1 = dim_list[sp_dim[ran1]]
				dim2 = dim_list[sp_dim[ran2]]
				[num1, num2, num3] = setPartition_1(num, 2)
				sp_dict = copy.deepcopy(sp_init)
				sp_dict[dim1] *= num1
				sp_dict[dim2] *= num2
				if sp_dict not in list:
					gen_num += 1
					list.append(sp_dict)
				
				iter_num +=1
				if iter_num > 1000:
					break
			return list
	elif sp_type == 2:
		#--- hybrid
		assert(len(sp_dim) >= 2)
		while gen_num < TH:
			id_list = list(range(len(sp_dim)))
			random.shuffle(id_list)
			dim1 = dim_list[sp_dim[id_list[0]]]
			dim2 = dim_list[sp_dim[id_list[1]]]
			[num1, num2, num3] = setPartition_1(num, 2)
			sp_dict = copy.deepcopy(sp_init)
			sp_dict[dim1] *= num1
			sp_dict[dim2] *= num2
			if sp_dict not in list and num1 != 1 and num2 != 1:
				gen_num += 1
				list.append(sp_dict)
			
			iter_num +=1
			if iter_num > 1000:
				break
		return list

def getSpatialParallel(HW_param, chiplet_parallel, PE_parallel, numTH = 20):

	spatial_parallel_init = {"pe":{"P":1, "Q":1, "C":1, "K":1, "R":1, "S":1}, "chiplet":{"P":1, "Q":1, "C":1, "K":1, "R":1, "S":1}, "package":{"P":1, "Q":1, "C":1, "K":1, "R":1, "S":1}}
	for dim in HW_param["intra_PE"]:
		spatial_parallel_init["pe"][dim] = HW_param["intra_PE"][dim]
	
	chiplet_H = HW_param["Chiplet"][0]
	chiplet_W = HW_param["Chiplet"][1]
	chiplet_num = chiplet_H * chiplet_W
	PE_H = HW_param["PE"][0]
	PE_W = HW_param["PE"][1]
	PE_num = PE_H * PE_W

	parallel_select, parallel_type = config_parallel_type(chiplet_parallel,PE_parallel)	
	chiplet_sp_dim = parallel_select["Chiplet"]
	chiplet_sp_type = parallel_type["Chiplet"]
	pe_sp_dim = parallel_select["PE"]
	pe_sp_type = parallel_type["PE"]

	chiplet_list = getSPPartitonList(chiplet_num, chiplet_sp_dim, chiplet_sp_type)
	TH_pe = int(numTH/len(chiplet_list))
	pe_list = getSPPartitonList(PE_num, pe_sp_dim, pe_sp_type, TH_pe)

	spatial_parallel_list = []

	for chiplet in chiplet_list:
		for pe in pe_list:
			spatial_parallel = copy.deepcopy(spatial_parallel_init)
			spatial_parallel["chiplet"] = pe
			spatial_parallel["package"] = chiplet
			spatial_parallel_list.append(spatial_parallel)
	
	return spatial_parallel_list


def gaTest_NoC(num_gen, num_iter, result_dir, save_all_records, record_dir, GaType, HW_param, memory_param, layer_dict, spatial_parallel_list, NoC_param, optimization_objective, layer_name_dict, architecture, multi_layer_tag="initial", if_multicast=1, io_die_tag = 1):
	
	edp_res_min_dict = {"mid":{} , "head":{}, "tail":{}}
	energy_min_dict = {"mid":{} , "head":{}, "tail":{}}
	delay_min_dict = {"mid":{} , "head":{}, "tail":{}}
	code_min_dict = {"mid":{} , "head":{}, "tail":{}}
	degrade_ratio_min_dict = {"mid":{} , "head":{}, "tail":{}}
	NoC_DR_dict = {"mid":{} , "head":{}, "tail":{}}
	L2_to_DRAM_DR_dict = {"mid":{} , "head":{}, "tail":{}}
	DRAM_to_L2_DR_dict = {"mid":{} , "head":{}, "tail":{}}
	NoP_DR_dict = {"mid":{} , "head":{}, "tail":{}}
	degrade_ratio_dict_min_dict = {"mid":{} , "head":{}, "tail":{}}
	input_DRAM_flit_needed_min_dict = {"mid":{} , "head":{}, "tail":{}}
	input_L2_flit_needed_min_dict = {"mid":{} , "head":{}, "tail":{}}
	weight_DRAM_flit_needed_min_dict = {"mid":{} , "head":{}, "tail":{}}
	weight_L2_flit_needed_min_dict = {"mid":{} , "head":{}, "tail":{}}
	output_rd_flit_needed_min_dict = {"mid":{} , "head":{}, "tail":{}}
	output_wr_flit_needed_min_dict = {"mid":{} , "head":{}, "tail":{}}
	chiplet_spatial_parallel_dict = {"mid":{} , "head":{}, "tail":{}}
	compuation_cycles_min_dict = {"mid":{} , "head":{}, "tail":{}}
	iter_num_dict = {"mid":{} , "head":{}, "tail":{}}
	edp_total = 0
	excel_data = []
	i_act_enough_dict = {}
	par_num_detail_dict = {}
	par_num_dict = {}

	layer_id = 0

	for layer_name in layer_dict:
		layer_id += 1
		# ---输出文件
		print("-------START LAYER : {} ----------".format(layer_name))
		if save_all_records == 1:
			record_filename = record_dir + "/" + layer_name + "_" + multi_layer_tag + ".xls"
		else:
			record_filename = None
		
		# --- 初始化参数
		network_param = layer_dict[layer_name]
		i_act_enough = layer_dict[layer_name]['i_act_enough']
		i_act_enough_dict[layer_name] = i_act_enough
		if "weight_enough" in layer_dict[layer_name]:
			weight_enough = layer_dict[layer_name]['weight_enough']
			par_num = layer_dict[layer_name]['P_par'] * layer_dict[layer_name]['Q_par']
			par_num_detail_dict[layer_name] = [layer_dict[layer_name]['P_par'], layer_dict[layer_name]['Q_par']]
		else:
			weight_enough = 0
			par_num = 1
			par_num_detail_dict[layer_name] = [1,1]
		par_num_dict[layer_name] = par_num

		GAGen = GaEncode(GaType, network_param, HW_param, debug=0)
		GA_Solver = GASolver(num_gen, num_iter, memory_param, NoC_param, if_multicast, architecture, record_filename, optimization_objective, i_act_enough,weight_enough, par_num, multi_layer_tag, io_die_tag = io_die_tag)
		GA_Solver.setGAGen(GAGen)

		# --- Initial: 初始进行硬件并行度方案的择优
		fitness_sp_dict = {}
		generation_dict = {}
		best_out_dict = {}
		record_dict = {}
		fitness_dict = {}
		loop_tile_dict_dict = {}
		for sp_id in range(len(spatial_parallel_list)):
			spatial_parallel = spatial_parallel_list[sp_id]
			GA_Solver.GAGen.setSpatialParallel(spatial_parallel)
			GA_Solver.total_reset()
			GA_Solver.getFirstGeneration()
			fitness_sp_dict[sp_id] = GA_Solver.best_out["fitness"]
			generation_dict[sp_id] = GA_Solver.generation
			best_out_dict[sp_id] = GA_Solver.best_out
			record_dict[sp_id] = GA_Solver.record
			fitness_dict[sp_id] = GA_Solver.fitness
			loop_tile_dict_dict[sp_id] = copy.deepcopy(GA_Solver.GAGen.loop_tile_dict)
		
		# --- --- 排序：去除fitness很差的硬件并行度方案，并且按照fitness数值设置迭代次数，fitness越小迭代次数越多
		# --- --- 通过设置max_TH的大小、TH的大小，可以对迭代次数进行增加或减少
		# print("spatial_parallel_list : ")
		# for sp in spatial_parallel_list:
		# 	print(sp)
		# print("fitness_sp_dict : ", fitness_sp_dict)
		fitness_dict_order = sorted(fitness_sp_dict.items(), key=lambda item:item[1])
		fitness_best = fitness_dict_order[0][1]
		TH_max = math.ceil(len(spatial_parallel_list) * 0.4)
		TH_min = 2
		max_TH = 1.5
		sp_select = []
		num_iter_list = {}
		iter_num_total = 0
		for i in range(len(fitness_dict_order)):
			sp_id = fitness_dict_order[i][0]
			fitness = fitness_dict_order[i][1]
			fitness_ratio = fitness/fitness_best
			if i < TH_min:
				sp_select.append(sp_id)
				num_iter_list[sp_id] = num_iter
				iter_num_total += num_iter_list[sp_id]
			elif i < TH_max and fitness_ratio < max_TH:
				sp_select.append(sp_id)
				num_iter_list[sp_id] = math.ceil(num_iter / (fitness_ratio*fitness_ratio))
				iter_num_total += num_iter_list[sp_id]
			else:
				break
		# print("fitness_dict_order : ", fitness_dict_order)
		
		# --- 遗传算法GA求解
		GA_Solver.total_reset()
		for sp_id in sp_select:
			spatial_parallel = spatial_parallel_list[sp_id]
			GA_Solver.GAGen.setSpatialParallel(spatial_parallel, loop_tile_dict_dict[sp_id])
			# print("SpatialParallel is {} ----------".format(spatial_parallel))
			GA_Solver.generation_reset()
			GA_Solver.generation = generation_dict[sp_id]
			GA_Solver.num_iter = num_iter_list[sp_id]
			GA_Solver.fitness = fitness_dict[sp_id]
			GA_Solver.gaIter(0)

		#for spatial_parallel in spatial_parallel_list:
		#	GA_Solver.GAGen.setSpatialParallel(spatial_parallel)
		#	print("SpatialParallel is {} ----------".format(spatial_parallel))
		#	GA_Solver.gaIter()
		
		if save_all_records == 1:
			GA_Solver.evaluationRecord()
		
		# --- 各层结果记录
		for pos in ["mid", "head", "tail"]:
			edp_res_min_dict[pos][layer_name] = GA_Solver.best_out["edp"][pos]
			energy_min_dict[pos][layer_name] = GA_Solver.best_out["e_sum"][pos]
			delay_min_dict[pos][layer_name] = GA_Solver.best_out["delay"][pos]
			code_min_dict[pos][layer_name] = GA_Solver.best_out["code"]
			degrade_ratio_min_dict[pos][layer_name] = GA_Solver.best_out["degrade_ratio"][pos]

			NoC_DR_dict[pos][layer_name] = GA_Solver.best_out["degrade_ratio_dict"][pos]["NoC"]
			L2_to_DRAM_DR_dict[pos][layer_name] = GA_Solver.best_out["degrade_ratio_dict"][pos]["L2_to_DRAM"]
			DRAM_to_L2_DR_dict[pos][layer_name] = GA_Solver.best_out["degrade_ratio_dict"][pos]["DRAM_to_L2"]
			NoP_DR_dict[pos][layer_name] = GA_Solver.best_out["degrade_ratio_dict"][pos]["NoP"]

			input_DRAM_flit_needed_min_dict[pos][layer_name] = GA_Solver.best_out["flit_needed_dict"][pos]["input_DRAM"]
			weight_DRAM_flit_needed_min_dict[pos][layer_name] = GA_Solver.best_out["flit_needed_dict"][pos]["weight_DRAM"]
			input_L2_flit_needed_min_dict[pos][layer_name] = GA_Solver.best_out["flit_needed_dict"][pos]["input_L2"]
			weight_L2_flit_needed_min_dict[pos][layer_name] = GA_Solver.best_out["flit_needed_dict"][pos]["weight_L2"]
			output_rd_flit_needed_min_dict[pos][layer_name] = GA_Solver.best_out["flit_needed_dict"][pos]["output_rd"]
			output_wr_flit_needed_min_dict[pos][layer_name] = GA_Solver.best_out["flit_needed_dict"][pos]["output_wr"]
			chiplet_spatial_parallel_dict[pos][layer_name] = GA_Solver.best_out["flit_needed_dict"][pos]["chiplet_parallel"]

			degrade_ratio_dict_min_dict[pos][layer_name] = GA_Solver.best_out["degrade_ratio_dict"][pos]
			compuation_cycles_min_dict[pos][layer_name] = GA_Solver.best_out["compuation_cycles"][pos]
			iter_num_dict[pos][layer_name] = iter_num_total


		edp_total += GA_Solver.best_out["edp"]["mid"]
		
		layer_data_record = [layer_name, edp_res_min_dict["mid"][layer_name], energy_min_dict["mid"][layer_name], delay_min_dict["mid"][layer_name], str(code_min_dict["mid"][layer_name]), degrade_ratio_min_dict["mid"][layer_name], NoC_DR_dict["mid"][layer_name], L2_to_DRAM_DR_dict["mid"][layer_name], DRAM_to_L2_DR_dict["mid"][layer_name], str(degrade_ratio_dict_min_dict["mid"][layer_name]), input_DRAM_flit_needed_min_dict["mid"][layer_name], input_L2_flit_needed_min_dict["mid"][layer_name], weight_DRAM_flit_needed_min_dict["mid"][layer_name], weight_L2_flit_needed_min_dict["mid"][layer_name], output_rd_flit_needed_min_dict["mid"][layer_name], output_wr_flit_needed_min_dict["mid"][layer_name], compuation_cycles_min_dict["mid"][layer_name], iter_num_dict["mid"][layer_name]]
		# excel_data.append(layer_data_record)
		# print('------- layer_data_record = ', layer_data_record)

	# --- 结果输出 mid
	file_1 = result_dir + "/final_result_record_" + multi_layer_tag + "_midLayer.txt"
	f = open(file_1,'w')
	print("edp: ", edp_res_min_dict["mid"], file=f)
	print("energy: ", energy_min_dict["mid"], file=f)
	print("delay: ", delay_min_dict["mid"], file=f)
	print("code: ", code_min_dict["mid"], file = f)
	print("degrade_ratio: ", degrade_ratio_min_dict["mid"], file = f)
	print("NoC_DR: ", NoC_DR_dict["mid"], file = f)
	print("L2_to_DRAM_DR: ", L2_to_DRAM_DR_dict["mid"], file = f)
	print("DRAM_to_L2_DR: ", DRAM_to_L2_DR_dict["mid"], file = f)
	print("NoP_DR: ", NoP_DR_dict["mid"], file = f)
	print("input_DRAM_flit_needed: ",  input_DRAM_flit_needed_min_dict["mid"], file = f)
	print("weight_DRAM_flit_needed: ",  weight_DRAM_flit_needed_min_dict["mid"], file = f)
	print("input_L2_flit_needed: ",  input_L2_flit_needed_min_dict["mid"], file = f)
	print("weight_L2_flit_needed: ",  weight_L2_flit_needed_min_dict["mid"], file = f)
	print("output_rd_flit_needed: ",  output_rd_flit_needed_min_dict["mid"], file = f)
	print("output_wr_flit_needed: ",  output_wr_flit_needed_min_dict["mid"], file = f)
	print("chiplet_spatial_parallel: ",  chiplet_spatial_parallel_dict["mid"], file = f)
	print("compuation_cycles: ", compuation_cycles_min_dict["mid"], file = f)
	print("iter_num: ", iter_num_dict["mid"], file = f)
	print("i_act_enough: ", i_act_enough_dict, file = f)
	print("edp_total: ", edp_total, file = f)
	print("par_num: ", par_num_dict, file = f)
	print("par_num_detail: ", par_num_detail_dict, file = f)
	print("layer_name_dict: ", layer_name_dict, file=f)
	f.close()

	# --- 结果输出 head
	file_2 = result_dir + "/final_result_record_" + multi_layer_tag + "_headLayer.txt"
	f = open(file_2,'w')
	print("edp: ", edp_res_min_dict["head"], file=f)
	print("energy: ", energy_min_dict["head"], file=f)
	print("delay: ", delay_min_dict["head"], file=f)
	print("code: ", code_min_dict["head"], file = f)
	print("degrade_ratio: ", degrade_ratio_min_dict["head"], file = f)
	print("NoC_DR: ", NoC_DR_dict["head"], file = f)
	print("L2_to_DRAM_DR: ", L2_to_DRAM_DR_dict["head"], file = f)
	print("DRAM_to_L2_DR: ", DRAM_to_L2_DR_dict["head"], file = f)
	print("NoP_DR: ", NoP_DR_dict["head"], file = f)
	print("input_DRAM_flit_needed: ",  input_DRAM_flit_needed_min_dict["head"], file = f)
	print("weight_DRAM_flit_needed: ",  weight_DRAM_flit_needed_min_dict["head"], file = f)
	print("input_L2_flit_needed: ",  input_L2_flit_needed_min_dict["head"], file = f)
	print("weight_L2_flit_needed: ",  weight_L2_flit_needed_min_dict["head"], file = f)
	print("output_rd_flit_needed: ",  output_rd_flit_needed_min_dict["head"], file = f)
	print("output_wr_flit_needed: ",  output_wr_flit_needed_min_dict["head"], file = f)
	print("chiplet_spatial_parallel: ",  chiplet_spatial_parallel_dict["head"], file = f)
	print("compuation_cycles: ", compuation_cycles_min_dict["head"], file = f)
	print("iter_num_dict: ", iter_num_dict["head"], file = f)
	print("i_act_enough_dict: ", i_act_enough_dict, file = f)
	print("edp_total: ", edp_total, file = f)
	print("par_num: ", par_num_dict, file = f)
	print("par_num_detail: ", par_num_detail_dict, file = f)
	print("layer_name_dict: ", layer_name_dict, file=f)
	f.close()

	# --- 结果输出 tail
	file_3 = result_dir + "/final_result_record_" + multi_layer_tag + "_tailLayer.txt"
	f = open(file_3,'w')
	print("edp: ", edp_res_min_dict["tail"], file=f)
	print("energy: ", energy_min_dict["tail"], file=f)
	print("delay: ", delay_min_dict["tail"], file=f)
	print("code: ", code_min_dict["tail"], file = f)
	print("degrade_ratio: ", degrade_ratio_min_dict["tail"], file = f)
	print("NoC_DR: ", NoC_DR_dict["tail"], file = f)
	print("L2_to_DRAM_DR: ", L2_to_DRAM_DR_dict["tail"], file = f)
	print("DRAM_to_L2_DR: ", DRAM_to_L2_DR_dict["tail"], file = f)
	print("NoP_DR: ", NoP_DR_dict["tail"], file = f)
	print("input_DRAM_flit_needed: ",  input_DRAM_flit_needed_min_dict["tail"], file = f)
	print("weight_DRAM_flit_needed: ",  weight_DRAM_flit_needed_min_dict["tail"], file = f)
	print("input_L2_flit_needed: ",  input_L2_flit_needed_min_dict["tail"], file = f)
	print("weight_L2_flit_needed: ",  weight_L2_flit_needed_min_dict["tail"], file = f)
	print("output_rd_flit_needed: ",  output_rd_flit_needed_min_dict["tail"], file = f)
	print("output_wr_flit_needed: ",  output_wr_flit_needed_min_dict["tail"], file = f)
	print("chiplet_spatial_parallel: ",  chiplet_spatial_parallel_dict["tail"], file = f)
	print("compuation_cycles: ", compuation_cycles_min_dict["tail"], file = f)
	print("iter_num_dict: ", iter_num_dict["tail"], file = f)
	print("i_act_enough_dict: ", i_act_enough_dict, file = f)
	print("edp_total: ", edp_total, file = f)
	print("par_num: ", par_num_dict, file = f)
	print("par_num_detail: ", par_num_detail_dict, file = f)
	print("layer_name_dict: ", layer_name_dict, file=f)
	f.close()

	# --- excel 结果输出
	workbook = openpyxl.Workbook()
	sheet = workbook.get_sheet_by_name('Sheet') 
	# 写入标题
	column_tite = ["layer_name","fitness","energy", "delay", "code", "degrade_ratio(DR)", "NoC_DR", "L2_to_DRAM_DR", "DRAM_to_L2_DR", "DR_dict", "input_DRAM_flit_needed", "input_L2_flit_needed", "weight_DRAM_flit_needed", "weight_L2_flit_needed", "output_rd_flit_needed", "output_wr_flit_needed", "computation_cycles", "iter_num"]
	for col,column in enumerate(column_tite):
		sheet.cell(1, col+1, column)
	
	# print('----- excel_data = ', excel_data)
	# 写入每一行
	for row, data in enumerate(excel_data):
		for col, column_data in enumerate(data):
			sheet.cell(row+2, col+1, column_data)
	filename = result_dir + '/final_result_record_' + multi_layer_tag + '.xls'
	workbook.save(filename)

if __name__ == '__main__':
	parser = argparse.ArgumentParser()
	parser.add_argument('--architecture', type=str, default="cmesh", help='hardware architecture type (cmesh, mesh, ring)')
	parser.add_argument('--app_name', type=str, default="resnet50", help='NN model name')
	# parser.add_argument('--alg', type=str, default="GA", help='algorithnm (GA, random)')				# random
	# parser.add_argument('--encode_type', type=str, default="index", help='encode type (index, num)')				# random
	parser.add_argument('--chiplet_num_max', type=int, default=None, help='1-16')
	parser.add_argument('--chiplet_num_min', type=int, default=None, help='1-16')
	# parser.add_argument('--chiplet_num', type=int, default=None, help='1-16')
	parser.add_argument('--chiplet_parallel', type=str, default="All", help='chiplet level spatial parallel type') # K_stable, P_stable, PK_stable, C_stable, KC_stable
	parser.add_argument('--PE_parallel', type=str, default="All", help='PE level spatial parallel type')
	parser.add_argument('--debug_open', type=int, default=0, help='debug mode (will print info)')
	parser.add_argument('--save_all_records', type=int, default=0, help='save all record')
	parser.add_argument('--layer_fuse_tag', type=int, default=1, help='layer_fuse_tag')
	parser.add_argument('--optimization_objective', type=str, default="edp", help='optimization_objective')
	opt = parser.parse_args()

	abs_path = os.path.dirname(os.path.abspath(__file__))
	architecture = opt.architecture
	app_name = opt.app_name
	alg = 'GA'
	encode_type = 'index'
	chiplet_num_max = opt.chiplet_num_max
	chiplet_num_min = opt.chiplet_num_min
	chiplet_parallel = opt.chiplet_parallel
	PE_parallel = opt.PE_parallel
	debug_open = opt.debug_open
	save_all_records = opt.save_all_records
	layer_fuse_tag = opt.layer_fuse_tag
	optimization_objective = opt.optimization_objective

	record_outdir = os.path.join(abs_path, "output_record")
	os.makedirs(record_outdir, exist_ok=True)
	record_outdir = os.path.join(record_outdir, architecture + "_" + app_name)
	os.makedirs(record_outdir, exist_ok=True)
	record_outdir = os.path.join(record_outdir, chiplet_parallel + "_and_" + PE_parallel)
	os.makedirs(record_outdir, exist_ok=True)
	record_outdir = os.path.join(record_outdir, alg + "_" + encode_type)
	os.makedirs(record_outdir, exist_ok=True)

	for i in range(chiplet_num_min-1, chiplet_num_max):
		chiplet_num = i+1
		result_outdir = os.path.join(abs_path, "result")
		os.makedirs(result_outdir, exist_ok=True)
		result_outdir = os.path.join(result_outdir, "intraLayer")
		os.makedirs(result_outdir, exist_ok=True)
		result_outdir = os.path.join(result_outdir, architecture + "_" + app_name)
		os.makedirs(result_outdir, exist_ok=True)
		result_outdir = os.path.join(result_outdir, "chiplet_num_"+str(chiplet_num))
		os.makedirs(result_outdir, exist_ok=True)
		result_outdir = os.path.join(result_outdir, alg + "_" + encode_type)
		os.makedirs(result_outdir, exist_ok=True)
		result_outdir = os.path.join(result_outdir, chiplet_parallel + "_and_" + PE_parallel)
		os.makedirs(result_outdir, exist_ok=True)

		# --- 硬件参数
		HW_param = {"Chiplet":[4,4],"PE":[4,4],"intra_PE":{"C":8,"K":8}}       	
		memory_param = {"OL1":3 ,"OL2":48,"AL1":8,"AL2":48,"WL1":32,"WL2":32}		# mesh mem
		noc_topo = 'Mesh'

		if architecture == "cmesh_DDR4" or architecture == "cmesh_HBM":
			HW_param["maxChiplet"] = [4, 4]
			io_die_tag = 1
		elif architecture == "mesh_DDR4" or architecture == "mesh_HBM":
			HW_param["maxChiplet"] = [4, 4]
			io_die_tag = 0
		elif architecture == "ring_DDR4" or architecture == "ring_HBM":
			HW_param["maxChiplet"] = [4, 2]
			io_die_tag = 0
		else:
			print("Error architecture(), not supported".format(architecture))
			raise NotImplementedError

		if chiplet_num != None:
			HW_param["Chiplet"] = [int(chiplet_num), 1]

		NoC_w = HW_param["PE"][1] + 1
		NOC_NODE_NUM = NoC_w * HW_param["PE"][0]
		NoP_w = HW_param["Chiplet"][1] + 1
		NOP_SIZE = NoP_w * HW_param["Chiplet"][0]
		TOPO_param = {"NoC_w":NoC_w, "NOC_NODE_NUM": NOC_NODE_NUM, "NoP_w": NoP_w, "NOP_SIZE": NOP_SIZE,"nop_scale_ratio": nop_bandwidth/noc_bandwidth}
		
		# --- 生成noc-nop结构图
		NoC_param, all_sim_node_num = construct_noc_nop_topo(TOPO_param["NOC_NODE_NUM"],TOPO_param["NoC_w"], TOPO_param["NOP_SIZE"],TOPO_param["NoP_w"], TOPO_param["nop_scale_ratio"], topology = noc_topo)
		if_multicast = 1

		# --- 神经网络参数
		layer_dict, layer_name_dict = getLayerParam(app_name)
		print('layer_dict = ', layer_dict)

		partiton_size_list = {"P":2, "Q":2}
		chiplet_num = HW_param["Chiplet"][0] * HW_param["Chiplet"][1]
		OL2_mem = memory_param["OL2"] * 8 * 1024 / act_wgt_width * chiplet_num
		WL2_mem = memory_param["WL2"] * 8 * 1024 / act_wgt_width * chiplet_num
		mem_size_dict = {"O" : OL2_mem, "W" : WL2_mem}
		layer_dict, head_layer_dict, tail_layer_dict = getLayerParamForMulti(layer_dict, mem_size_dict)
		print("layer_name_dict----------------------------------")
		print(layer_name_dict)
		print(" ")
		print("layer_dict---------------------------------------")
		for layer_name, layer in layer_dict.items():
			print(layer_name, ": ", layer)
		print(" ")
		print("head_layer_dict----------------------------------")
		for layer_name, layer in head_layer_dict.items():
			print(layer_name, ": ", layer)
		print(" ")
		print("tail_layer_dict----------------------------------")
		for layer_name, layer in tail_layer_dict.items():
			print(layer_name, ": ", layer)
		assert(len(head_layer_dict) == len(tail_layer_dict))
		# --- 获得空间并行度
		spatial_parallel_list = getSpatialParallel(HW_param, chiplet_parallel, PE_parallel)
		
		# --- 迭代参数
		num_gen = 50
		# num_iter = int(20 / 8 * min(chiplet_num, 8))
		num_iter = 20

		iterTime = num_gen * num_iter

		gaTest_NoC(num_gen, num_iter, result_outdir, save_all_records, record_outdir, encode_type, HW_param, memory_param, layer_dict, spatial_parallel_list, NoC_param, optimization_objective, layer_name_dict, architecture, io_die_tag=io_die_tag)
		if len(head_layer_dict) > 0:
			gaTest_NoC(num_gen, num_iter, result_outdir, save_all_records, record_outdir, encode_type, HW_param, memory_param, head_layer_dict, spatial_parallel_list, NoC_param, optimization_objective, layer_name_dict, architecture, multi_layer_tag = "headFuse", io_die_tag=io_die_tag)
			gaTest_NoC(num_gen, num_iter, result_outdir, save_all_records, record_outdir, encode_type, HW_param, memory_param, tail_layer_dict, spatial_parallel_list, NoC_param, optimization_objective, layer_name_dict, architecture, multi_layer_tag = "tailFuse", io_die_tag=io_die_tag)


